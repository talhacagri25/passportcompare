"""Passport fraud detection — single-file production pipeline.

Scans a directory of passport images, computes a face embedding per passport,
and outputs ranked candidate fraud pairs (same person under different identities)
for human review.

Pipeline:
    1. SCRFD (InsightFace) face detection + 5-keypoint extraction
    2. ArcFace template alignment → 112×112 face crop
    3. AdaFace IR-101 (CVLface, WebFace12M) → 512-d L2-normalized embedding
    4. FAISS IndexFlatIP (cosine = inner product on unit vectors)
    5. Top-K nearest neighbours per passport, threshold-filtered candidates

Models are bundled with this repository, except the 250 MB AdaFace weights file
which lives as a GitHub release asset:

    SCRFD detector:    models/buffalo_l/det_10g.onnx                (~17 MB, in repo)
    AdaFace IR-101:    checkpoints/adaface_ir101_webface12m.pt      (~250 MB, release asset)

Download the AdaFace weights once on the prod server (one-time, ~250 MB):

    mkdir -p checkpoints
    wget -O checkpoints/adaface_ir101_webface12m.pt \
      https://github.com/talhacagri25/passportcompare/releases/download/v1.0/adaface_ir101_webface12m.pt

After that, no internet is required at runtime.

Usage:
    pip install -r requirements.txt
    python passport_fraud.py                          # uses defaults
    python passport_fraud.py --input /path/to/dir --threshold 0.20
    python passport_fraud.py --workers 16 --batch-size 256 --fp16   # A100 prod

Speed knobs (default tuned for an RTX 4070 Super-class GPU):
    --workers       N CPU worker processes for image decode + SCRFD detection.
                    Scale up to ~min(CPU cores - 2, 16) for big A100 boxes.
    --batch-size    AdaFace inference batch size. 128 is conservative; on A100
                    80 GB you can push 256-1024. Each sample is ~30 MB activation.
    --fp16          Run AdaFace in FP16 autocast on GPU (A100/H100/4090 only).
                    ~2× faster, no measurable accuracy loss for inference.

Outputs (under --output):
    candidates.xlsx          Excel: rank, cosine, file names, paths (sortable)
    candidates.csv           CSV mirror of the same data
    candidates.html          Visual review panel (open in browser)
    fraud_candidates/        Copies of suspect passport images for inspection.
                             Original input files are NEVER modified — copies are
                             converted to JPG and renamed <rank>_<cosine>_<pid>.jpg
                             so the file browser sorts them by suspicion.
    summary.txt              Run statistics (totals, score distribution)
    embeddings.npy           (N, 512) float32, L2-normalized — for re-use
    passport_ids.txt         passport ID per row of embeddings.npy
    failed_detection.txt     IDs where SCRFD found no face
    index.faiss              FAISS index for re-querying without re-embedding

Threshold guidance:
    Defaults to 0.20. The right operating point depends on the real distribution
    of passport quality. We recommend collecting a small labelled sample
    (~200-500 pairs of known same-person and known different-person) and
    sweeping thresholds in [0.10, 0.45] to pick where precision/recall balance
    matches your operational tolerance for false positives.

The threshold is the only knob that needs prod calibration; the model and
pipeline are pre-trained on millions of faces and do not require fine-tuning.
"""
from __future__ import annotations

import argparse
import csv
import os
import sys
import time
import traceback
from collections import namedtuple
from dataclasses import dataclass
from pathlib import Path

import numpy as np
from PIL import Image
from tqdm import tqdm

import torch
import torch.nn.functional as F
from torch import nn
from torch.nn import (BatchNorm1d, BatchNorm2d, Conv2d, Dropout, Linear,
                      MaxPool2d, Module, PReLU, Sequential)


# ════════════════════════════════════════════════════════════════════════════
# AdaFace IR-101 architecture (vendored from CVLface, MIT)
# ════════════════════════════════════════════════════════════════════════════

class _Flatten(Module):
    def forward(self, x: torch.Tensor) -> torch.Tensor:
        return x.reshape(x.size(0), -1)


class _BasicBlockIR(Module):
    def __init__(self, in_channel: int, depth: int, stride: int):
        super().__init__()
        if in_channel == depth:
            self.shortcut_layer = MaxPool2d(1, stride)
        else:
            self.shortcut_layer = Sequential(
                Conv2d(in_channel, depth, (1, 1), stride, bias=False),
                BatchNorm2d(depth),
            )
        self.res_layer = Sequential(
            BatchNorm2d(in_channel),
            Conv2d(in_channel, depth, (3, 3), (1, 1), 1, bias=False),
            BatchNorm2d(depth),
            PReLU(depth),
            Conv2d(depth, depth, (3, 3), stride, 1, bias=False),
            BatchNorm2d(depth),
        )

    def forward(self, x: torch.Tensor) -> torch.Tensor:
        return self.res_layer(x) + self.shortcut_layer(x)


_Bottleneck = namedtuple("_Bottleneck", ["in_channel", "depth", "stride"])


def _block(in_channel: int, depth: int, num_units: int, stride: int = 2):
    return [_Bottleneck(in_channel, depth, stride)] + \
           [_Bottleneck(depth, depth, 1) for _ in range(num_units - 1)]


def _blocks_ir101():
    return [
        _block(64, 64, 3),
        _block(64, 128, 13),
        _block(128, 256, 30),
        _block(256, 512, 3),
    ]


class _IR101Backbone(Module):
    """IR-101 backbone for 112×112 input → 512-d face embedding."""

    def __init__(self, output_dim: int = 512):
        super().__init__()
        self.input_layer = Sequential(
            Conv2d(3, 64, (3, 3), 1, 1, bias=False),
            BatchNorm2d(64),
            PReLU(64),
        )
        modules = []
        for block in _blocks_ir101():
            for unit in block:
                modules.append(_BasicBlockIR(unit.in_channel, unit.depth, unit.stride))
        self.body = Sequential(*modules)
        self.output_layer = Sequential(
            BatchNorm2d(512),
            Dropout(0.4),
            _Flatten(),
            Linear(512 * 7 * 7, output_dim),
            BatchNorm1d(output_dim, affine=False),
        )

    def forward(self, x: torch.Tensor) -> torch.Tensor:
        x = self.input_layer(x)
        x = self.body(x)
        return self.output_layer(x)


# ════════════════════════════════════════════════════════════════════════════
# Weight loading — strictly local, no network
# ════════════════════════════════════════════════════════════════════════════

_ADAFACE_RELEASE_URL = (
    "https://github.com/talhacagri25/passportcompare/releases/download/"
    "v1.0/adaface_ir101_webface12m.pt"
)


def _verify_bundled(path: Path, label: str, min_size_mb: int,
                    download_url: str | None = None) -> Path:
    if not path.exists():
        msg = f"{label} not found at {path}."
        if download_url:
            msg += (f"\n\nDownload the file once with:\n"
                    f"    mkdir -p {path.parent}\n"
                    f"    wget -O {path} \\\n"
                    f"        {download_url}\n")
        raise FileNotFoundError(msg)
    size_mb = path.stat().st_size / (1024 * 1024)
    if size_mb < min_size_mb:
        raise ValueError(
            f"{label} at {path} looks truncated ({size_mb:.1f} MB; "
            f"expected ≥{min_size_mb} MB). Re-download the file."
        )
    return path


def load_adaface_ir101(weights_path: Path, device: torch.device) -> _IR101Backbone:
    state_dict = torch.load(weights_path, map_location="cpu", weights_only=True)
    if any(k.startswith("net.") for k in state_dict):
        state_dict = {k[len("net."):]: v for k, v in state_dict.items()
                      if k.startswith("net.")}
    model = _IR101Backbone(output_dim=512)
    missing, unexpected = model.load_state_dict(state_dict, strict=False)
    if missing or unexpected:
        print(f"[adaface] load: {len(missing)} missing, {len(unexpected)} unexpected keys")
    model.eval()
    model.to(device)
    return model


# ════════════════════════════════════════════════════════════════════════════
# Detection + embedding pipeline (multi-worker, batched, optional FP16)
# ════════════════════════════════════════════════════════════════════════════

# A100-class GPUs leave most of their VRAM idle if we run images one-by-one.
# Architecture:
#   - DataLoader with N CPU workers, each running its own SCRFD ONNX session
#     (CPU EP, threadsafe, no GPU contention). Each __getitem__ returns the
#     aligned 112×112 face crop as uint8 — or None if no face was detected.
#   - Main thread pulls batches of crops, transfers them to GPU as one tensor,
#     and runs AdaFace IR-101 in a single big batch (optional FP16 autocast).
#
# This pattern keeps the GPU busy while CPUs decode/detect in parallel.


class _PassportDataset(torch.utils.data.Dataset):
    """Per-image: PIL decode → SCRFD detect → ArcFace template align → 112×112 crop.

    Each worker lazily creates its own SCRFD ONNX session (CPU EP) and reuses it
    for every image it processes. Returning None on failure keeps the
    main loop's accounting simple.
    """

    def __init__(self, paths: list[Path], root: Path, scrfd_path: str,
                 det_size: tuple[int, int], det_threshold: float):
        self.paths = paths
        self.root = root
        self.scrfd_path = scrfd_path
        self.det_size = det_size
        self.det_threshold = det_threshold
        self._det = None
        self._face_align = None

    def _lazy_init(self) -> None:
        if self._det is not None:
            return
        from insightface.model_zoo.scrfd import SCRFD
        from insightface.utils import face_align
        self._det = SCRFD(model_file=self.scrfd_path)
        # ctx_id=-1 forces CPU EP — workers must NOT touch CUDA (forked processes
        # share no CUDA context with the main process)
        self._det.prepare(ctx_id=-1, input_size=self.det_size)
        self._face_align = face_align

    def __len__(self) -> int:
        return len(self.paths)

    def __getitem__(self, idx: int):
        self._lazy_init()
        path = self.paths[idx]
        pid = passport_id_from_path(path, self.root)
        img = _read_image_robust(path)
        if img is None:
            return pid, None
        try:
            bboxes, kpss = self._det.detect(img[:, :, ::-1], metric="default")
        except Exception:
            return pid, None
        if bboxes is None or len(bboxes) == 0:
            return pid, None
        scores = bboxes[:, 4]
        sizes = (bboxes[:, 2] - bboxes[:, 0]) * (bboxes[:, 3] - bboxes[:, 1])
        valid = scores >= self.det_threshold
        if not valid.any():
            return pid, None
        idxs = np.where(valid)[0]
        best = idxs[int(np.argmax(sizes[valid]))]
        kps = kpss[best]
        aligned_bgr = self._face_align.norm_crop(img[:, :, ::-1],
                                                  landmark=kps, image_size=112)
        # Return as RGB uint8 (B,H,W,3); main thread will permute and normalize on GPU
        return pid, aligned_bgr[:, :, ::-1].copy()


def _collate(batch):
    """List[(pid, face|None)] → (pids, faces_array|None, valid_mask)."""
    pids = [b[0] for b in batch]
    faces = [b[1] for b in batch]
    valid_mask = np.array([f is not None for f in faces], dtype=bool)
    if not valid_mask.any():
        return pids, None, valid_mask
    valid_faces = [f for f in faces if f is not None]
    arr = np.stack(valid_faces, axis=0)  # (B, 112, 112, 3) uint8
    return pids, arr, valid_mask


class FaceEmbedder:
    """Holds AdaFace IR-101 on GPU + the SCRFD model path workers will use.

    The detection step runs in DataLoader workers (CPU EP); the embedding step
    runs on the main process's CUDA device with whatever batch size fits.
    """

    def __init__(
        self,
        adaface_weights_path: Path,
        scrfd_model_path: Path,
        det_size: tuple[int, int] = (640, 640),
        det_score_threshold: float = 0.4,
        device: str = "cuda",
        fp16: bool = False,
    ):
        self.device = torch.device(device if torch.cuda.is_available() else "cpu")
        self.det_threshold = det_score_threshold
        self.det_size = det_size
        self.fp16 = fp16 and self.device.type == "cuda"

        _verify_bundled(scrfd_model_path, "SCRFD det_10g.onnx", min_size_mb=10)
        _verify_bundled(adaface_weights_path, "AdaFace IR-101 .pt",
                        min_size_mb=200, download_url=_ADAFACE_RELEASE_URL)
        self.scrfd_path = str(scrfd_model_path)

        self.model = load_adaface_ir101(adaface_weights_path, self.device)

        # Warm up CUDA kernels with a single forward pass — first call is slow
        # because of cuBLAS/cuDNN autotuning
        if self.device.type == "cuda":
            with torch.no_grad():
                dummy = torch.zeros(1, 3, 112, 112, device=self.device)
                if self.fp16:
                    with torch.amp.autocast("cuda", dtype=torch.float16):
                        self.model(dummy)
                else:
                    self.model(dummy)
            torch.cuda.synchronize()

    def embed_batch(self, faces_uint8: np.ndarray) -> np.ndarray:
        """(B, 112, 112, 3) uint8 RGB → (B, 512) float32, L2-normalized."""
        x = torch.from_numpy(faces_uint8).to(self.device, non_blocking=True)
        x = x.permute(0, 3, 1, 2).float() / 255.0
        x = (x - 0.5) / 0.5
        with torch.no_grad():
            if self.fp16:
                with torch.amp.autocast("cuda", dtype=torch.float16):
                    emb = self.model(x)
            else:
                emb = self.model(x)
            emb = F.normalize(emb.float(), p=2, dim=1)
        return emb.cpu().numpy().astype(np.float32)


# ════════════════════════════════════════════════════════════════════════════
# I/O + index + candidates
# ════════════════════════════════════════════════════════════════════════════

IMAGE_SUFFIXES = {".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff", ".webp"}


def _read_image_robust(path: Path) -> np.ndarray | None:
    """Try PIL, then OpenCV — handles many TIF variants (LZW, JPEG-in-TIFF,
    multi-page, 16-bit grayscale) that PIL alone refuses to decode.

    Returns RGB uint8 (H, W, 3) or None if every backend fails.
    """
    # PIL — fastest for standard formats
    try:
        return np.array(Image.open(path).convert("RGB"))
    except Exception:
        pass
    # OpenCV — handles JPEG-in-TIFF, LZW, packbits, etc. that PIL chokes on
    try:
        import cv2
        bgr = cv2.imread(str(path), cv2.IMREAD_COLOR)
        if bgr is not None:
            return bgr[:, :, ::-1].copy()
    except Exception:
        pass
    # Multi-page TIF: try first page explicitly via PIL's seek
    try:
        with Image.open(path) as im:
            im.seek(0)
            return np.array(im.convert("RGB"))
    except Exception:
        pass
    return None


def discover_passports(root: Path) -> list[Path]:
    """Recursively collect all images under root, sorted for determinism."""
    paths = [p for p in root.rglob("*") if p.suffix.lower() in IMAGE_SUFFIXES]
    paths.sort()
    return paths


def passport_id_from_path(path: Path, root: Path) -> str:
    """Use the path relative to root, slash-flattened, as the passport ID.

    e.g. /input/2024/A1234.jpg under /input → "2024__A1234"
    Avoids collisions when filenames repeat across subdirs.
    """
    rel = path.relative_to(root).with_suffix("")
    return str(rel).replace(os.sep, "__")


def extract_all(
    paths: list[Path],
    root: Path,
    embedder: FaceEmbedder,
    out_dir: Path,
    workers: int = 8,
    batch_size: int = 128,
) -> tuple[np.ndarray, list[str]]:
    """Multi-worker extraction. Workers do PIL+SCRFD on CPU; main GPU-batches AdaFace.

    Workers run SCRFD on CPU EP (the ONNX model is small, 17 MB, fast on CPU and
    parallelizes cleanly across processes). The main thread aggregates aligned
    crops, ships them to GPU as a single tensor per batch, and runs AdaFace
    once per batch.
    """
    dataset = _PassportDataset(
        paths, root, embedder.scrfd_path, embedder.det_size, embedder.det_threshold,
    )
    loader = torch.utils.data.DataLoader(
        dataset,
        batch_size=batch_size,
        num_workers=workers,
        collate_fn=_collate,
        pin_memory=False,                  # we do our own .to(device) below
        persistent_workers=workers > 0,
        prefetch_factor=2 if workers > 0 else None,
    )

    embeddings: list[np.ndarray] = []
    pids_ok: list[str] = []
    failed: list[str] = []
    t0 = time.time()
    n_total = len(paths)

    pbar = tqdm(total=n_total, desc="embedding", unit="img")
    for pids_batch, faces_arr, valid_mask in loader:
        for pid, ok in zip(pids_batch, valid_mask):
            if not ok:
                failed.append(pid)
        if faces_arr is not None:
            emb = embedder.embed_batch(faces_arr)
            valid_pids = [pid for pid, ok in zip(pids_batch, valid_mask) if ok]
            for j, pid in enumerate(valid_pids):
                embeddings.append(emb[j])
                pids_ok.append(pid)
        pbar.update(len(pids_batch))
    pbar.close()

    elapsed = time.time() - t0
    rate = n_total / max(0.001, elapsed)
    print(f"[embed] {len(embeddings)} OK / {len(failed)} failed in {elapsed:.1f}s "
          f"({elapsed / max(1, n_total) * 1000:.1f} ms/img, {rate:.0f} img/s)")

    emb_arr = (np.stack(embeddings, 0).astype(np.float32)
               if embeddings else np.zeros((0, 512), np.float32))
    np.save(out_dir / "embeddings.npy", emb_arr)
    (out_dir / "passport_ids.txt").write_text("\n".join(pids_ok))
    (out_dir / "failed_detection.txt").write_text("\n".join(failed))
    return emb_arr, pids_ok


def build_index_and_query(emb: np.ndarray, top_k: int, out_dir: Path):
    """Build FAISS IndexFlatIP and self-query top-K (excluding self)."""
    import faiss
    n, d = emb.shape
    # enforce unit norm
    norms = np.linalg.norm(emb, axis=1)
    if not np.allclose(norms, 1.0, atol=1e-3):
        emb = emb / np.maximum(norms, 1e-12)[:, None]
    index = faiss.IndexFlatIP(d)
    index.add(emb)
    print(f"[index] built: ntotal={index.ntotal}, d={d}")

    k_plus_self = min(top_k + 1, n)
    t0 = time.time()
    scores, indices = index.search(emb, k_plus_self)
    print(f"[index] top-{top_k} self-query in {time.time() - t0:.2f}s "
          f"({(time.time() - t0) / max(1, n) * 1000:.2f} ms/q)")

    # Drop self-hit (column 0 for unique vectors)
    if (indices[:, 0] == np.arange(n)).mean() > 0.99:
        indices = indices[:, 1:]
        scores = scores[:, 1:]
    else:
        rows_idx, rows_sc = [], []
        for i in range(n):
            mask = indices[i] != i
            rows_idx.append(indices[i][mask][:top_k])
            rows_sc.append(scores[i][mask][:top_k])
        indices = np.stack(rows_idx)
        scores = np.stack(rows_sc)

    faiss.write_index(index, str(out_dir / "index.faiss"))
    return indices, scores


def collect_candidates(
    pids: list[str],
    indices: np.ndarray,
    scores: np.ndarray,
    threshold: float,
) -> list[tuple[tuple[str, str], float]]:
    """Deduplicated undirected pairs above threshold, sorted by cosine desc."""
    seen: dict[tuple[str, str], float] = {}
    n, k = indices.shape
    for i in range(n):
        a = pids[i]
        for j in range(k):
            sc = float(scores[i, j])
            if sc < threshold:
                continue
            b = pids[indices[i, j]]
            if a == b:
                continue
            key = (a, b) if a < b else (b, a)
            if key not in seen or sc > seen[key]:
                seen[key] = sc
    return sorted(seen.items(), key=lambda kv: -kv[1])


def copy_candidate_images(
    candidates: list[tuple[tuple[str, str], float]],
    image_paths: dict[str, Path],
    review_dir: Path,
) -> list[dict]:
    """Copy passport images of every candidate pair into review_dir as JPGs.

    Original files are read-only. Filenames embed rank and cosine so a file
    browser sorts them by suspicion (highest cosine first).
    """
    review_dir.mkdir(parents=True, exist_ok=True)
    rows = []
    for rank, ((a, b), sc) in enumerate(tqdm(candidates, desc="copying suspect images"), 1):
        pa = image_paths.get(a)
        pb = image_paths.get(b)
        if pa is None or pb is None:
            continue
        out_a = review_dir / f"{rank:04d}_{sc:.3f}_{a}.jpg"
        out_b = review_dir / f"{rank:04d}_{sc:.3f}_{b}.jpg"
        try:
            Image.open(pa).convert("RGB").save(out_a, quality=92)
            Image.open(pb).convert("RGB").save(out_b, quality=92)
        except Exception as exc:
            print(f"[warn] could not copy pair {a}/{b}: {exc}")
            continue
        rows.append({
            "rank": rank,
            "cosine": round(sc, 4),
            "passport_a": a,
            "passport_b": b,
            "review_image_a": str(out_a.relative_to(review_dir.parent)),
            "review_image_b": str(out_b.relative_to(review_dir.parent)),
            "original_a": str(pa),
            "original_b": str(pb),
        })
    return rows


def write_candidates_csv(rows: list[dict], out_path: Path) -> None:
    fieldnames = ["rank", "cosine", "passport_a", "passport_b",
                  "review_image_a", "review_image_b", "original_a", "original_b"]
    with out_path.open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(rows)
    print(f"[out] candidates.csv -> {out_path} ({len(rows)} pairs)")


def write_candidates_xlsx(rows: list[dict], summary: dict, out_path: Path) -> None:
    """Excel with two sheets: 'Fraud Candidates' (sortable) + 'Summary'."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        print("[warn] openpyxl not installed — skipping xlsx output. "
              "Install with: pip install openpyxl")
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Fraud Candidates"
    headers = ["rank", "cosine", "passport_a", "passport_b",
               "review_image_a", "review_image_b", "original_a", "original_b"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
    # Auto-width approximations
    widths = {"A": 6, "B": 9, "C": 16, "D": 16, "E": 36, "F": 36, "G": 60, "H": 60}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Summary")
    ws2.append(["metric", "value"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for k, v in summary.items():
        ws2.append([k, v])
    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 24

    wb.save(out_path)
    print(f"[out] candidates.xlsx -> {out_path}")


# ════════════════════════════════════════════════════════════════════════════
# Optional HTML report
# ════════════════════════════════════════════════════════════════════════════

_HTML_HEAD = """<!doctype html>
<html><head><meta charset="utf-8"><title>Passport fraud candidates</title>
<style>
body{font-family:system-ui,sans-serif;max-width:1200px;margin:24px auto;padding:0 16px;}
h1{font-size:20px}
.row{display:flex;gap:16px;align-items:center;border-bottom:1px solid #eee;padding:12px 0}
.row img{height:160px;border:1px solid #ccc}
.score{font-size:14px;color:#444;min-width:100px}
.cosine{font-size:18px;font-weight:bold;color:#a13}
.ids{font-family:monospace;font-size:13px}
.summary{background:#f6f6f8;padding:10px 14px;border-radius:6px;margin-bottom:16px}
</style></head><body>
"""


def render_html(rows: list[dict], threshold: float, out_path: Path) -> None:
    """Side-by-side passport pair viewer using the COPIED review images.

    Uses relative paths so the HTML stays valid when the output dir is moved.
    """
    out_path.parent.mkdir(parents=True, exist_ok=True)
    body = [_HTML_HEAD]
    body.append("<h1>Passport fraud candidates</h1>")
    body.append(f"<div class='summary'>{len(rows)} candidate pairs at cosine ≥ "
                f"{threshold}. Sorted by similarity descending. "
                "Higher cosine = more suspicious.</div>")
    for r in rows:
        body.append("<div class='row'>")
        body.append(f"<div class='score'><span class='cosine'>{r['cosine']:.3f}</span><br>"
                    f"<span class='ids'>#{r['rank']}</span></div>")
        body.append(f"<div><img src='{r['review_image_a']}'><br>"
                    f"<span class='ids'>{r['passport_a']}</span></div>")
        body.append(f"<div><img src='{r['review_image_b']}'><br>"
                    f"<span class='ids'>{r['passport_b']}</span></div>")
        body.append("</div>")
    body.append("</body></html>")
    out_path.write_text("".join(body))
    print(f"[out] candidates.html -> {out_path}")


# ════════════════════════════════════════════════════════════════════════════
# Main
# ════════════════════════════════════════════════════════════════════════════

DEFAULT_INPUT = "/TeftisDataScience/BerkayDeneme/Pasaport_Resim_Kontrol/Indirilen_Pasaport"
DEFAULT_OUTPUT = "/TeftisDataScience/BerkayDeneme/TalhaOutput"

# Bundled model paths — relative to this script's location
_SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_ADAFACE = str(_SCRIPT_DIR / "checkpoints/adaface_ir101_webface12m.pt")
DEFAULT_SCRFD = str(_SCRIPT_DIR / "models/buffalo_l/det_10g.onnx")


def build_summary(
    n_input: int,
    n_embedded: int,
    n_failed: int,
    candidates: list[tuple[tuple[str, str], float]],
    threshold: float,
    capped_at: int | None = None,
) -> dict:
    summary = {
        "input_images_found": n_input,
        "successfully_embedded": n_embedded,
        "failed_detection": n_failed,
        "detection_rate": round(n_embedded / max(1, n_input), 4),
        "fraud_threshold": threshold,
        "candidate_pairs": len(candidates),
    }
    if capped_at is not None:
        summary["candidate_pairs_capped_at"] = capped_at
    if candidates:
        cosines = np.array([sc for _, sc in candidates], dtype=np.float32)
        summary.update({
            "candidate_cosine_max": round(float(cosines.max()), 4),
            "candidate_cosine_min": round(float(cosines.min()), 4),
            "candidate_cosine_median": round(float(np.median(cosines)), 4),
            "candidate_cosine_mean": round(float(cosines.mean()), 4),
        })
        # Severity buckets
        for lo, hi in [(0.7, 1.01), (0.5, 0.7), (0.4, 0.5),
                        (0.3, 0.4), (0.2, 0.3)]:
            n = int(((cosines >= lo) & (cosines < hi)).sum())
            summary[f"pairs_in_cosine_[{lo:.2f},{hi:.2f})"] = n
    return summary


def main() -> int:
    p = argparse.ArgumentParser(
        description="Passport fraud detection — single-file pipeline",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument("--input", type=Path, default=Path(DEFAULT_INPUT),
                   help="Directory of passport images (recursive)")
    p.add_argument("--output", type=Path, default=Path(DEFAULT_OUTPUT),
                   help="Output directory for embeddings and candidates")
    p.add_argument("--adaface-weights", type=Path, default=Path(DEFAULT_ADAFACE),
                   help="AdaFace IR-101 .pt path (bundled with repo)")
    p.add_argument("--scrfd-model", type=Path, default=Path(DEFAULT_SCRFD),
                   help="SCRFD det_10g.onnx path (bundled with repo)")
    p.add_argument("--threshold", type=float, default=0.30,
                   help="Cosine threshold for fraud candidates. At 40k+ scale, 0.20 "
                        "produces unactionable noise (10⁵+ pairs). Start with 0.30, "
                        "raise to 0.35-0.40 if the candidate list is still too large.")
    p.add_argument("--top-k", type=int, default=20,
                   help="Top-K nearest neighbours per passport")
    p.add_argument("--max-candidates", type=int, default=2000,
                   help="Cap on candidate pairs written + image-copied. The full "
                        "above-threshold set is logged for traceability, but only "
                        "the top --max-candidates by cosine are materialized as "
                        "image copies / Excel rows.")
    p.add_argument("--device", default="cuda",
                   help="cuda or cpu")
    p.add_argument("--det-size", type=int, default=640,
                   help="SCRFD detection input size (square)")
    p.add_argument("--workers", type=int, default=8,
                   help="DataLoader worker processes for PIL decode + SCRFD detection")
    p.add_argument("--batch-size", type=int, default=128,
                   help="AdaFace embedding batch size on GPU")
    p.add_argument("--fp16", action="store_true",
                   help="Run AdaFace in FP16 autocast on GPU (A100/H100: ~2× faster)")
    p.add_argument("--no-html", action="store_true",
                   help="Skip HTML viewer rendering (default: render it)")
    p.add_argument("--no-resume", action="store_true",
                   help="Force re-extraction even if embeddings.npy exists")
    args = p.parse_args()

    if not args.input.exists():
        print(f"[err] input directory does not exist: {args.input}")
        return 2
    args.output.mkdir(parents=True, exist_ok=True)

    print(f"[info] input  = {args.input}")
    print(f"[info] output = {args.output}")
    print(f"[info] device = {args.device} (CUDA available: {torch.cuda.is_available()})")

    paths = discover_passports(args.input)
    print(f"[scan] found {len(paths)} images")
    if not paths:
        return 1
    image_paths = {passport_id_from_path(p, args.input): p for p in paths}

    # Resume if embeddings exist for the same passport set
    emb_path = args.output / "embeddings.npy"
    pids_path = args.output / "passport_ids.txt"
    if (not args.no_resume and emb_path.exists() and pids_path.exists()):
        cached_pids = pids_path.read_text().splitlines()
        if set(cached_pids) == set(image_paths) and len(cached_pids) > 0:
            print(f"[resume] using cached embeddings ({len(cached_pids)} passports)")
            emb = np.load(emb_path)
            pids = cached_pids
        else:
            cached_pids = None
    else:
        cached_pids = None

    if cached_pids is None:
        print(f"[info] workers={args.workers} batch_size={args.batch_size} "
              f"fp16={args.fp16}")
        embedder = FaceEmbedder(
            adaface_weights_path=args.adaface_weights,
            scrfd_model_path=args.scrfd_model,
            det_size=(args.det_size, args.det_size),
            device=args.device,
            fp16=args.fp16,
        )
        emb, pids = extract_all(
            paths, args.input, embedder, args.output,
            workers=args.workers, batch_size=args.batch_size,
        )

    if emb.shape[0] < 2:
        print(f"[warn] only {emb.shape[0]} usable embedding(s); cannot pair-search")
        return 0

    indices, scores = build_index_and_query(emb, args.top_k, args.output)

    candidates_full = collect_candidates(pids, indices, scores, args.threshold)
    print(f"[candidates] {len(candidates_full)} pairs above cosine {args.threshold}")
    if len(candidates_full) > args.max_candidates:
        print(f"[candidates] capping at top {args.max_candidates} by cosine "
              f"(use --max-candidates to change)")
    candidates = candidates_full[: args.max_candidates]

    review_dir = args.output / "fraud_candidates"
    rows = copy_candidate_images(candidates, image_paths, review_dir)
    print(f"[out] {len(rows)} suspect-pair image copies -> {review_dir}")

    write_candidates_csv(rows, args.output / "candidates.csv")
    summary = build_summary(
        len(paths), len(pids), len(paths) - len(pids),
        candidates_full, args.threshold,
        capped_at=args.max_candidates if len(candidates_full) > args.max_candidates else None,
    )
    write_candidates_xlsx(rows, summary, args.output / "candidates.xlsx")

    summary_txt = "\n".join(f"{k}: {v}" for k, v in summary.items())
    (args.output / "summary.txt").write_text(summary_txt + "\n")
    print(f"[out] summary.txt -> {args.output / 'summary.txt'}")

    if rows and not args.no_html:
        render_html(rows, args.threshold, args.output / "candidates.html")

    print()
    print("=" * 60)
    print(f"DONE — {len(rows)} candidate fraud pairs above cosine {args.threshold}")
    print(f"  Browse:  {args.output / 'candidates.html'}")
    print(f"  Excel:   {args.output / 'candidates.xlsx'}")
    print(f"  Images:  {review_dir}/")
    print("=" * 60)
    return 0


if __name__ == "__main__":
    sys.exit(main())
